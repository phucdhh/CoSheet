#!/usr/bin/env python3
"""Create a test XLSX file with multiple sheets"""
try:
    from openpyxl import Workbook
    
    wb = Workbook()
    
    # Sheet 1: Sales Data
    ws1 = wb.active
    ws1.title = "Sales Data"
    ws1['A1'] = "Product"
    ws1['B1'] = "Quantity"
    ws1['C1'] = "Price"
    ws1['A2'] = "Laptop"
    ws1['B2'] = 10
    ws1['C2'] = 1000
    ws1['A3'] = "Mouse"
    ws1['B3'] = 50
    ws1['C3'] = 25
    
    # Sheet 2: Employees
    ws2 = wb.create_sheet("Employees")
    ws2['A1'] = "Name"
    ws2['B1'] = "Department"
    ws2['C1'] = "Salary"
    ws2['A2'] = "John"
    ws2['B2'] = "IT"
    ws2['C2'] = 50000
    ws2['A3'] = "Jane"
    ws2['B3'] = "HR"
    ws2['C3'] = 45000
    
    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    ws3['A1'] = "Report"
    ws3['B1'] = "Q1 2024"
    ws3['A2'] = "Total Sales"
    ws3['B2'] = 1250
    ws3['A3'] = "Total Employees"
    ws3['B3'] = 2
    
    wb.save("/root/ethercalc/chrome_test/test_multisheet.xlsx")
    print("✅ Created test_multisheet.xlsx with 3 sheets")
    
except ImportError:
    print("❌ openpyxl not installed, installing...")
    import subprocess
    subprocess.check_call(['pip3', 'install', 'openpyxl'])
    print("✅ Installed openpyxl, please run script again")
